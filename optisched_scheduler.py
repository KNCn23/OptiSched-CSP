import json
import pandas as pd
from ortools.sat.python import cp_model
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

class OptiSchedSolver:
    def __init__(self, data_file):
        self.data = self._load_data(data_file)
        self.model = cp_model.CpModel()
        self.assignments = {}
        self.result_data = []

    def _load_data(self, data_file):
        with open(data_file, 'r', encoding='utf-8') as f:
            return json.load(f)

    def build_model(self):
        courses = self.data['courses']
        classrooms = self.data['classrooms']
        days = self.data['days']
        timeslots = self.data['timeslots']
        lecturers = {l['id']: l['name'] for l in self.data['lecturers']}

        for course in courses:
            for d_idx in range(len(days)):
                for slot in timeslots:
                    for room in classrooms:
                        self.assignments[(course['id'], d_idx, slot['id'], room['id'])] = self.model.NewBoolVar(
                            f'c{course["id"]}_d{d_idx}_s{slot["id"]}_r{room["id"]}'
                        )

        # Constraints
        for course in courses:
            self.model.Add(sum(self.assignments[(course['id'], d_idx, slot['id'], room['id'])]
                               for d_idx in range(len(days)) for slot in timeslots for room in classrooms) == course['hours_per_week'])

        for d_idx in range(len(days)):
            for slot in timeslots:
                for room in classrooms:
                    self.model.Add(sum(self.assignments[(course['id'], d_idx, slot['id'], room['id'])] for course in courses) <= 1)

        for d_idx in range(len(days)):
            for slot in timeslots:
                for l_id in lecturers:
                    l_courses = [c['id'] for c in courses if c['lecturer_id'] == l_id]
                    if l_courses:
                        self.model.Add(sum(self.assignments[(c_id, d_idx, slot['id'], room['id'])] for c_id in l_courses for room in classrooms) <= 1)

        for d_idx in range(len(days)):
            for slot in timeslots:
                if slot.get('is_lunch'):
                    for course in courses:
                        for room in classrooms:
                            self.model.Add(self.assignments[(course['id'], d_idx, slot['id'], room['id'])] == 0)

        for course in courses:
            for room in classrooms:
                if course['student_count'] > room['capacity']:
                    for d_idx in range(len(days)):
                        for slot in timeslots:
                            self.model.Add(self.assignments[(course['id'], d_idx, slot['id'], room['id'])] == 0)

    def solve(self, output_excel='ders_programi.xlsx'):
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = 30.0
        status = solver.Solve(self.model)
        
        if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
            self._collect_results(solver)
            self._export_to_excel(output_excel)
        else:
            print("❌ Çözüm bulunamadı!")

    def _collect_results(self, solver):
        days = self.data['days']
        timeslots = self.data['timeslots']
        classrooms = self.data['classrooms']
        courses = self.data['courses']
        lecturers = {l['id']: l['name'] for l in self.data['lecturers']}

        for d_idx, day in enumerate(days):
            for slot in timeslots:
                for room in classrooms:
                    for course in courses:
                        if solver.Value(self.assignments[(course['id'], d_idx, slot['id'], room['id'])]):
                            self.result_data.append({
                                "Gün": day,
                                "Saat": slot['time'],
                                "Sınıf": room['id'],
                                "Bilgi": f"{course['name']}\n{lecturers[course['lecturer_id']]}"
                            })
                if slot.get('is_lunch'):
                    self.result_data.append({"Gün": day, "Saat": slot['time'], "Sınıf": "LUNCH", "Bilgi": "--- ÖĞLE ARASI ---"})

    def _export_to_excel(self, filename):
        df = pd.DataFrame(self.result_data)
        days_order = self.data['days']
        slots_order = [s['time'] for s in self.data['timeslots']]

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            for room in self.data['classrooms']:
                room_df = df[(df['Sınıf'] == room['id']) | (df['Sınıf'] == "LUNCH")]
                if not room_df.empty:
                    pivot = room_df.pivot_table(index='Saat', columns='Gün', values='Bilgi', aggfunc='first')
                    pivot = pivot.reindex(columns=days_order, index=slots_order)
                    sheet_name = f'Sınıf {room["id"]}'
                    pivot.to_excel(writer, sheet_name=sheet_name)
                    
                    # STYLING
                    ws = writer.sheets[sheet_name]
                    self._apply_styles(ws)

            # TÜM OKUL ÖZET
            all_pivot = df.pivot_table(index=['Saat', 'Sınıf'], columns='Gün', values='Bilgi', aggfunc='first').reindex(columns=days_order)
            all_pivot.to_excel(writer, sheet_name='Tüm Okul Özet')
            self._apply_styles(writer.sheets['Tüm Okul Özet'])

        print(f"✅ Görsel ve Kutucuklu Excel Hazır: {filename}")

    def _apply_styles(self, ws):
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
        lunch_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                
                # Header Styling
                if cell.row == 1 or cell.column == 1:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                
                # Lunch Break Styling
                if cell.value == "--- ÖĞLE ARASI ---":
                    cell.fill = lunch_fill
                    cell.font = Font(italic=True, color="808080")

        # Column Widths
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20
        # Row Heights
        for row_dim in ws.row_dimensions.values():
            row_dim.height = 45

if __name__ == '__main__':
    scheduler = OptiSchedSolver('optisched_data.json')
    scheduler.build_model()
    scheduler.solve()
